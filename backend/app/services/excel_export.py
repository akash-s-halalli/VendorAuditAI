"""Excel export service for generating formatted XLSX reports."""

import io
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.document import Document
from app.models.finding import Finding
from app.models.remediation import RemediationTask
from app.models.vendor import Vendor
from app.schemas.export import (
    FindingExportFilters,
    RemediationExportFilters,
    VendorExportFilters,
)


# Excel style constants
SEVERITY_COLORS = {
    "critical": "FF0000",  # Red
    "high": "FFA500",  # Orange
    "medium": "FFFF00",  # Yellow
    "low": "00FF00",  # Green
    "info": "87CEEB",  # Light Blue
}

PRIORITY_COLORS = {
    "critical": "FF0000",  # Red
    "high": "FFA500",  # Orange
    "medium": "FFFF00",  # Yellow
    "low": "00FF00",  # Green
}

STATUS_COLORS = {
    "open": "FF6B6B",  # Light Red
    "acknowledged": "FFE066",  # Yellow
    "remediated": "69DB7C",  # Light Green
    "accepted": "74C0FC",  # Light Blue
    "false_positive": "E9ECEF",  # Gray
}

HEADER_FILL = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


async def get_vendors_for_export(
    db: AsyncSession,
    org_id: str,
    filters: VendorExportFilters | None = None,
) -> list[Vendor]:
    """Fetch vendors with filters for export."""
    query = select(Vendor).where(Vendor.organization_id == org_id)

    if filters:
        if filters.status:
            query = query.where(Vendor.status == filters.status)
        if filters.tier:
            query = query.where(Vendor.tier == filters.tier)
        if filters.category:
            query = query.where(Vendor.category == filters.category)
        if filters.contract_expiry_before:
            query = query.where(Vendor.contract_expiry <= filters.contract_expiry_before)
        if filters.contract_expiry_after:
            query = query.where(Vendor.contract_expiry >= filters.contract_expiry_after)
        if filters.assessment_due_before:
            query = query.where(
                Vendor.next_assessment_due <= filters.assessment_due_before
            )

    query = query.order_by(Vendor.name)
    result = await db.execute(query)
    return list(result.scalars().all())


async def get_findings_with_filters(
    db: AsyncSession,
    org_id: str,
    filters: FindingExportFilters | None = None,
) -> list[Finding]:
    """Fetch findings with extended filters for export."""
    query = (
        select(Finding)
        .join(Document, Finding.document_id == Document.id)
        .join(Vendor, Document.vendor_id == Vendor.id)
        .where(Vendor.organization_id == org_id)
    )

    if filters:
        if filters.status:
            query = query.where(Finding.status == filters.status)
        if filters.severity:
            query = query.where(Finding.severity == filters.severity)
        if filters.framework:
            query = query.where(Finding.framework == filters.framework)
        if filters.document_id:
            query = query.where(Finding.document_id == filters.document_id)
        if filters.vendor_id:
            query = query.where(Vendor.id == filters.vendor_id)
        if filters.created_from:
            query = query.where(
                func.date(Finding.created_at) >= filters.created_from
            )
        if filters.created_to:
            query = query.where(func.date(Finding.created_at) <= filters.created_to)

    query = query.order_by(Finding.created_at.desc())
    result = await db.execute(query)
    return list(result.scalars().all())


async def get_remediation_tasks_for_export(
    db: AsyncSession,
    org_id: str,
    filters: RemediationExportFilters | None = None,
) -> list[RemediationTask]:
    """Fetch remediation tasks with filters for export."""
    query = (
        select(RemediationTask)
        .join(Vendor, RemediationTask.vendor_id == Vendor.id)
        .options(
            selectinload(RemediationTask.finding),
            selectinload(RemediationTask.vendor),
            selectinload(RemediationTask.assignee),
        )
        .where(Vendor.organization_id == org_id)
    )

    if filters:
        if filters.status:
            query = query.where(RemediationTask.status == filters.status)
        if filters.priority:
            query = query.where(RemediationTask.priority == filters.priority)
        if filters.vendor_id:
            query = query.where(RemediationTask.vendor_id == filters.vendor_id)
        if filters.assignee_id:
            query = query.where(RemediationTask.assignee_id == filters.assignee_id)
        if filters.sla_breached is not None:
            query = query.where(RemediationTask.sla_breached == filters.sla_breached)
        if filters.due_before:
            query = query.where(RemediationTask.due_date <= filters.due_before)
        if filters.due_after:
            query = query.where(RemediationTask.due_date >= filters.due_after)
        if filters.overdue_only:
            query = query.where(
                RemediationTask.due_date < datetime.now(UTC),
                RemediationTask.status.not_in(
                    ["closed", "verified", "exception_approved"]
                ),
            )

    query = query.order_by(RemediationTask.due_date.asc().nullslast())
    result = await db.execute(query)
    return list(result.scalars().all())


async def get_vendor_documents(
    db: AsyncSession,
    vendor_id: str,
) -> list[Document]:
    """Fetch documents for a vendor."""
    query = (
        select(Document)
        .where(Document.vendor_id == vendor_id)
        .order_by(Document.created_at.desc())
    )
    result = await db.execute(query)
    return list(result.scalars().all())


async def get_vendor_findings(
    db: AsyncSession,
    vendor_id: str,
) -> list[Finding]:
    """Fetch findings for a vendor through their documents."""
    query = (
        select(Finding)
        .join(Document, Finding.document_id == Document.id)
        .where(Document.vendor_id == vendor_id)
        .order_by(Finding.created_at.desc())
    )
    result = await db.execute(query)
    return list(result.scalars().all())


async def get_vendor_remediation_tasks(
    db: AsyncSession,
    vendor_id: str,
) -> list[RemediationTask]:
    """Fetch remediation tasks for a vendor."""
    query = (
        select(RemediationTask)
        .options(
            selectinload(RemediationTask.finding),
            selectinload(RemediationTask.assignee),
        )
        .where(RemediationTask.vendor_id == vendor_id)
        .order_by(RemediationTask.due_date.asc().nullslast())
    )
    result = await db.execute(query)
    return list(result.scalars().all())


class ExcelStyler:
    """Utility class for applying Excel styles."""

    @staticmethod
    def apply_header_style(ws: Worksheet, row: int = 1) -> None:
        """Apply header style to the first row."""
        for cell in ws[row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

    @staticmethod
    def apply_severity_color(cell: Any, severity: str) -> None:
        """Apply color based on severity level."""
        color = SEVERITY_COLORS.get(severity.lower(), "FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # Use dark font for light backgrounds
        if severity.lower() in ["medium", "low", "info"]:
            cell.font = Font(color="000000")
        else:
            cell.font = Font(color="FFFFFF", bold=True)

    @staticmethod
    def apply_priority_color(cell: Any, priority: str) -> None:
        """Apply color based on priority level."""
        color = PRIORITY_COLORS.get(priority.lower(), "FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        if priority.lower() in ["medium", "low"]:
            cell.font = Font(color="000000")
        else:
            cell.font = Font(color="FFFFFF", bold=True)

    @staticmethod
    def apply_status_color(cell: Any, status: str) -> None:
        """Apply color based on status."""
        color = STATUS_COLORS.get(status.lower(), "FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(color="000000")

    @staticmethod
    def auto_size_columns(ws: Worksheet) -> None:
        """Auto-size columns based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    lines = cell_value.split("\n")
                    cell_length = max(len(line) for line in lines)
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            # Cap the width at 50 characters
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    @staticmethod
    def freeze_header_row(ws: Worksheet) -> None:
        """Freeze the header row."""
        ws.freeze_panes = "A2"


class VendorExcelExporter:
    """Generates Excel exports of vendor data with formatting."""

    COLUMNS = [
        "ID",
        "Name",
        "Description",
        "Website",
        "Tier",
        "Status",
        "Category",
        "Criticality Score",
        "Data Classification",
        "Contract Expiry",
        "Last Assessed",
        "Next Assessment Due",
        "Created At",
    ]

    def __init__(self, vendors: list[Vendor]):
        """Initialize the Excel exporter."""
        self.vendors = vendors

    def generate(self) -> bytes:
        """Generate Excel content from vendors."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendors"

        # Write header
        for col, header in enumerate(self.COLUMNS, 1):
            ws.cell(row=1, column=col, value=header)

        # Apply header styles
        ExcelStyler.apply_header_style(ws)

        # Write data rows
        for row_idx, vendor in enumerate(self.vendors, 2):
            ws.cell(row=row_idx, column=1, value=vendor.id)
            ws.cell(row=row_idx, column=2, value=vendor.name)
            ws.cell(row=row_idx, column=3, value=vendor.description or "")
            ws.cell(row=row_idx, column=4, value=vendor.website or "")

            # Tier cell with color
            tier_cell = ws.cell(row=row_idx, column=5, value=vendor.tier.upper() if vendor.tier else "")
            if vendor.tier:
                ExcelStyler.apply_priority_color(tier_cell, vendor.tier)

            ws.cell(row=row_idx, column=6, value=vendor.status or "")
            ws.cell(row=row_idx, column=7, value=vendor.category or "")
            ws.cell(row=row_idx, column=8, value=vendor.criticality_score)
            ws.cell(row=row_idx, column=9, value=vendor.data_classification or "")
            ws.cell(
                row=row_idx,
                column=10,
                value=(
                    vendor.contract_expiry.isoformat()
                    if vendor.contract_expiry
                    else ""
                ),
            )
            ws.cell(
                row=row_idx,
                column=11,
                value=(
                    vendor.last_assessed.strftime("%Y-%m-%d %H:%M")
                    if vendor.last_assessed
                    else ""
                ),
            )
            ws.cell(
                row=row_idx,
                column=12,
                value=(
                    vendor.next_assessment_due.isoformat()
                    if vendor.next_assessment_due
                    else ""
                ),
            )
            ws.cell(
                row=row_idx,
                column=13,
                value=(
                    vendor.created_at.strftime("%Y-%m-%d %H:%M")
                    if vendor.created_at
                    else ""
                ),
            )

            # Apply border to all cells in row
            for col in range(1, len(self.COLUMNS) + 1):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER
                ws.cell(row=row_idx, column=col).alignment = CELL_ALIGNMENT

        # Auto-size columns and freeze header
        ExcelStyler.auto_size_columns(ws)
        ExcelStyler.freeze_header_row(ws)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class FindingExcelExporter:
    """Generates Excel exports of findings with severity coloring."""

    COLUMNS = [
        "ID",
        "Title",
        "Description",
        "Severity",
        "Status",
        "Framework",
        "Control",
        "Evidence",
        "Remediation",
        "Impact",
        "Page",
        "Section",
        "Confidence",
        "Document ID",
        "Created At",
        "Resolved At",
    ]

    def __init__(self, findings: list[Finding]):
        """Initialize the Excel exporter."""
        self.findings = findings

    def generate(self) -> bytes:
        """Generate Excel content from findings."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Findings"

        # Write header
        for col, header in enumerate(self.COLUMNS, 1):
            ws.cell(row=1, column=col, value=header)

        # Apply header styles
        ExcelStyler.apply_header_style(ws)

        # Write data rows
        for row_idx, finding in enumerate(self.findings, 2):
            ws.cell(row=row_idx, column=1, value=finding.id)
            ws.cell(row=row_idx, column=2, value=finding.title)
            ws.cell(row=row_idx, column=3, value=finding.description)

            # Severity cell with color
            severity_cell = ws.cell(
                row=row_idx, column=4, value=finding.severity.upper() if finding.severity else ""
            )
            if finding.severity:
                ExcelStyler.apply_severity_color(severity_cell, finding.severity)

            # Status cell with color
            status_cell = ws.cell(row=row_idx, column=5, value=finding.status.upper() if finding.status else "")
            if finding.status:
                ExcelStyler.apply_status_color(status_cell, finding.status)

            ws.cell(row=row_idx, column=6, value=finding.framework or "")
            ws.cell(row=row_idx, column=7, value=finding.framework_control or "")
            ws.cell(row=row_idx, column=8, value=finding.evidence or "")
            ws.cell(row=row_idx, column=9, value=finding.remediation or "")
            ws.cell(row=row_idx, column=10, value=finding.impact or "")
            ws.cell(row=row_idx, column=11, value=finding.page_number)
            ws.cell(row=row_idx, column=12, value=finding.section_header or "")
            ws.cell(
                row=row_idx,
                column=13,
                value=(
                    f"{finding.confidence_score:.2f}"
                    if finding.confidence_score
                    else ""
                ),
            )
            ws.cell(row=row_idx, column=14, value=finding.document_id)
            ws.cell(
                row=row_idx,
                column=15,
                value=(
                    finding.created_at.strftime("%Y-%m-%d %H:%M")
                    if finding.created_at
                    else ""
                ),
            )
            ws.cell(
                row=row_idx,
                column=16,
                value=(
                    finding.resolved_at.strftime("%Y-%m-%d %H:%M")
                    if finding.resolved_at
                    else ""
                ),
            )

            # Apply border and alignment to all cells in row
            for col in range(1, len(self.COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGNMENT

        # Auto-size columns and freeze header
        ExcelStyler.auto_size_columns(ws)
        ExcelStyler.freeze_header_row(ws)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class RemediationExcelExporter:
    """Generates Excel exports of remediation tasks with priority coloring."""

    COLUMNS = [
        "ID",
        "Title",
        "Description",
        "Status",
        "Priority",
        "Vendor",
        "Assignee",
        "Finding ID",
        "Finding Title",
        "Due Date",
        "SLA Days",
        "SLA Breached",
        "Resolution Notes",
        "Created At",
        "Resolved At",
    ]

    def __init__(self, tasks: list[RemediationTask]):
        """Initialize the Excel exporter."""
        self.tasks = tasks

    def generate(self) -> bytes:
        """Generate Excel content from remediation tasks."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Remediation Tasks"

        # Write header
        for col, header in enumerate(self.COLUMNS, 1):
            ws.cell(row=1, column=col, value=header)

        # Apply header styles
        ExcelStyler.apply_header_style(ws)

        # Write data rows
        for row_idx, task in enumerate(self.tasks, 2):
            ws.cell(row=row_idx, column=1, value=task.id)
            ws.cell(row=row_idx, column=2, value=task.title)
            ws.cell(row=row_idx, column=3, value=task.description or "")
            ws.cell(row=row_idx, column=4, value=task.status.upper() if task.status else "")

            # Priority cell with color
            priority_cell = ws.cell(
                row=row_idx, column=5, value=task.priority.upper() if task.priority else ""
            )
            if task.priority:
                ExcelStyler.apply_priority_color(priority_cell, task.priority)

            ws.cell(row=row_idx, column=6, value=task.vendor.name if task.vendor else "")
            ws.cell(
                row=row_idx, column=7, value=task.assignee.email if task.assignee else ""
            )
            ws.cell(row=row_idx, column=8, value=task.finding_id or "")
            ws.cell(
                row=row_idx, column=9, value=task.finding.title if task.finding else ""
            )
            ws.cell(
                row=row_idx,
                column=10,
                value=task.due_date.strftime("%Y-%m-%d") if task.due_date else "",
            )
            ws.cell(row=row_idx, column=11, value=task.sla_days)

            # SLA breached cell with color
            sla_cell = ws.cell(
                row=row_idx, column=12, value="Yes" if task.sla_breached else "No"
            )
            if task.sla_breached:
                sla_cell.fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
                sla_cell.font = Font(color="FFFFFF", bold=True)

            ws.cell(row=row_idx, column=13, value=task.resolution_notes or "")
            ws.cell(
                row=row_idx,
                column=14,
                value=(
                    task.created_at.strftime("%Y-%m-%d %H:%M")
                    if task.created_at
                    else ""
                ),
            )
            ws.cell(
                row=row_idx,
                column=15,
                value=(
                    task.resolved_at.strftime("%Y-%m-%d %H:%M")
                    if task.resolved_at
                    else ""
                ),
            )

            # Apply border and alignment to all cells in row
            for col in range(1, len(self.COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGNMENT

        # Auto-size columns and freeze header
        ExcelStyler.auto_size_columns(ws)
        ExcelStyler.freeze_header_row(ws)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class ComplianceReportExporter:
    """Generates comprehensive compliance reports in Excel format with database access."""

    def __init__(
        self,
        vendor_id: str,
        db: AsyncSession,
        frameworks: list[str] | None = None,
        include_findings: bool = True,
        include_remediation: bool = True,
        include_summary: bool = True,
    ):
        """Initialize the compliance report exporter."""
        self.vendor_id = vendor_id
        self.db = db
        self.frameworks = frameworks
        self.include_findings = include_findings
        self.include_remediation = include_remediation
        self.include_summary = include_summary
        self.vendor: Vendor | None = None
        self.findings: list[Finding] = []
        self.remediation_tasks: list[RemediationTask] = []

    async def _load_data(self) -> None:
        """Load vendor data from database."""
        # Load vendor
        query = select(Vendor).where(Vendor.id == self.vendor_id)
        result = await self.db.execute(query)
        self.vendor = result.scalar_one_or_none()

        if not self.vendor:
            return

        # Load findings
        findings_query = (
            select(Finding)
            .join(Document, Finding.document_id == Document.id)
            .where(Document.vendor_id == self.vendor_id)
        )
        if self.frameworks:
            findings_query = findings_query.where(Finding.framework.in_(self.frameworks))
        findings_query = findings_query.order_by(Finding.created_at.desc())
        result = await self.db.execute(findings_query)
        self.findings = list(result.scalars().all())

        # Load remediation tasks
        tasks_query = (
            select(RemediationTask)
            .options(
                selectinload(RemediationTask.finding),
                selectinload(RemediationTask.assignee),
            )
            .where(RemediationTask.vendor_id == self.vendor_id)
            .order_by(RemediationTask.due_date.asc().nullslast())
        )
        result = await self.db.execute(tasks_query)
        self.remediation_tasks = list(result.scalars().all())

    def _add_summary_sheet(self, wb: Workbook) -> None:
        """Add executive summary sheet."""
        ws = wb.create_sheet("Executive Summary", 0)

        # Title
        ws.merge_cells("A1:D1")
        title_cell = ws.cell(row=1, column=1, value="Compliance Report")
        title_cell.font = Font(bold=True, size=18, color="1A365D")
        title_cell.alignment = Alignment(horizontal="center")

        # Vendor info
        ws.cell(row=3, column=1, value="Vendor:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=self.vendor.name if self.vendor else "Unknown")
        ws.cell(row=4, column=1, value="Tier:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=self.vendor.tier.upper() if self.vendor and self.vendor.tier else "N/A")
        ws.cell(row=5, column=1, value="Status:").font = Font(bold=True)
        ws.cell(row=5, column=2, value=self.vendor.status.upper() if self.vendor and self.vendor.status else "N/A")
        ws.cell(row=6, column=1, value="Report Date:").font = Font(bold=True)
        ws.cell(row=6, column=2, value=datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC"))

        # Findings summary
        ws.cell(row=8, column=1, value="Findings Summary").font = Font(bold=True, size=14)

        # Count by severity
        severity_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0, "info": 0}
        for finding in self.findings:
            sev = finding.severity.lower() if finding.severity else "info"
            if sev in severity_counts:
                severity_counts[sev] += 1

        summary_headers = ["Severity", "Count"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        row_idx = 10
        for severity, count in severity_counts.items():
            ws.cell(row=row_idx, column=1, value=severity.upper())
            cell = ws.cell(row=row_idx, column=2, value=count)
            cell.border = THIN_BORDER
            ws.cell(row=row_idx, column=1).border = THIN_BORDER
            row_idx += 1

        ws.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=len(self.findings)).font = Font(bold=True)
        ws.cell(row=row_idx, column=2).border = THIN_BORDER

        # Remediation summary
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="Remediation Summary").font = Font(
            bold=True, size=14
        )

        row_idx += 1
        ws.cell(row=row_idx, column=1, value="Total Tasks:").font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=len(self.remediation_tasks))

        open_tasks = sum(
            1 for t in self.remediation_tasks if t.status not in ["closed", "verified"]
        )
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="Open Tasks:").font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=open_tasks)

        breached_tasks = sum(1 for t in self.remediation_tasks if t.sla_breached)
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="SLA Breached:").font = Font(bold=True)
        cell = ws.cell(row=row_idx, column=2, value=breached_tasks)
        if breached_tasks > 0:
            cell.font = Font(color="FF0000", bold=True)

        ExcelStyler.auto_size_columns(ws)

    def _add_findings_sheet(self, wb: Workbook) -> None:
        """Add findings detail sheet."""
        ws = wb.create_sheet("Findings")

        columns = [
            "Title",
            "Severity",
            "Status",
            "Framework",
            "Control",
            "Description",
            "Remediation",
            "Impact",
            "Page",
            "Confidence",
        ]

        # Write header
        for col, header in enumerate(columns, 1):
            ws.cell(row=1, column=col, value=header)

        ExcelStyler.apply_header_style(ws)

        # Write data
        for row_idx, finding in enumerate(self.findings, 2):
            ws.cell(row=row_idx, column=1, value=finding.title)

            severity_cell = ws.cell(
                row=row_idx, column=2, value=finding.severity.upper() if finding.severity else ""
            )
            if finding.severity:
                ExcelStyler.apply_severity_color(severity_cell, finding.severity)

            status_cell = ws.cell(row=row_idx, column=3, value=finding.status.upper() if finding.status else "")
            if finding.status:
                ExcelStyler.apply_status_color(status_cell, finding.status)

            ws.cell(row=row_idx, column=4, value=finding.framework or "")
            ws.cell(row=row_idx, column=5, value=finding.framework_control or "")
            ws.cell(row=row_idx, column=6, value=finding.description)
            ws.cell(row=row_idx, column=7, value=finding.remediation or "")
            ws.cell(row=row_idx, column=8, value=finding.impact or "")
            ws.cell(row=row_idx, column=9, value=finding.page_number)
            ws.cell(
                row=row_idx,
                column=10,
                value=(
                    f"{finding.confidence_score:.2f}"
                    if finding.confidence_score
                    else ""
                ),
            )

            for col in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGNMENT

        ExcelStyler.auto_size_columns(ws)
        ExcelStyler.freeze_header_row(ws)

    def _add_remediation_sheet(self, wb: Workbook) -> None:
        """Add remediation tasks sheet."""
        ws = wb.create_sheet("Remediation Tasks")

        columns = [
            "Title",
            "Status",
            "Priority",
            "Assignee",
            "Due Date",
            "SLA Breached",
            "Finding",
            "Resolution Notes",
        ]

        # Write header
        for col, header in enumerate(columns, 1):
            ws.cell(row=1, column=col, value=header)

        ExcelStyler.apply_header_style(ws)

        # Write data
        for row_idx, task in enumerate(self.remediation_tasks, 2):
            ws.cell(row=row_idx, column=1, value=task.title)
            ws.cell(row=row_idx, column=2, value=task.status.upper() if task.status else "")

            priority_cell = ws.cell(
                row=row_idx, column=3, value=task.priority.upper() if task.priority else ""
            )
            if task.priority:
                ExcelStyler.apply_priority_color(priority_cell, task.priority)

            ws.cell(
                row=row_idx, column=4, value=task.assignee.email if task.assignee else ""
            )
            ws.cell(
                row=row_idx,
                column=5,
                value=task.due_date.strftime("%Y-%m-%d") if task.due_date else "",
            )

            sla_cell = ws.cell(
                row=row_idx, column=6, value="Yes" if task.sla_breached else "No"
            )
            if task.sla_breached:
                sla_cell.fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
                sla_cell.font = Font(color="FFFFFF", bold=True)

            ws.cell(row=row_idx, column=7, value=task.finding.title if task.finding else "")
            ws.cell(row=row_idx, column=8, value=task.resolution_notes or "")

            for col in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGNMENT

        ExcelStyler.auto_size_columns(ws)
        ExcelStyler.freeze_header_row(ws)

    async def generate(self) -> bytes:
        """Generate comprehensive compliance report."""
        # Load data from database
        await self._load_data()

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Add sheets based on configuration
        if self.include_summary:
            self._add_summary_sheet(wb)
        if self.include_findings:
            self._add_findings_sheet(wb)
        if self.include_remediation:
            self._add_remediation_sheet(wb)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class VendorReportExporter:
    """Generates comprehensive vendor report in Excel format with database access."""

    def __init__(
        self,
        vendor_id: str,
        db: AsyncSession,
    ):
        """Initialize the vendor report exporter."""
        self.vendor_id = vendor_id
        self.db = db
        self.vendor: Vendor | None = None
        self.documents: list[Document] = []
        self.findings: list[Finding] = []
        self.remediation_tasks: list[RemediationTask] = []

    async def _load_data(self) -> None:
        """Load vendor data from database."""
        # Load vendor
        query = select(Vendor).where(Vendor.id == self.vendor_id)
        result = await self.db.execute(query)
        self.vendor = result.scalar_one_or_none()

        if not self.vendor:
            return

        # Load documents
        self.documents = await get_vendor_documents(self.db, self.vendor_id)

        # Load findings
        self.findings = await get_vendor_findings(self.db, self.vendor_id)

        # Load remediation tasks
        self.remediation_tasks = await get_vendor_remediation_tasks(self.db, self.vendor_id)

    def _add_overview_sheet(self, wb: Workbook) -> None:
        """Add vendor overview sheet."""
        ws = wb.create_sheet("Vendor Overview", 0)

        # Title
        ws.merge_cells("A1:D1")
        title_cell = ws.cell(row=1, column=1, value=f"Vendor Report: {self.vendor.name if self.vendor else 'Unknown'}")
        title_cell.font = Font(bold=True, size=18, color="1A365D")
        title_cell.alignment = Alignment(horizontal="center")

        if not self.vendor:
            ws.cell(row=3, column=1, value="Vendor not found")
            return

        # Basic info
        info_data = [
            ("Vendor ID:", self.vendor.id),
            ("Name:", self.vendor.name),
            ("Description:", self.vendor.description or "N/A"),
            ("Website:", self.vendor.website or "N/A"),
            ("Tier:", self.vendor.tier.upper() if self.vendor.tier else "N/A"),
            ("Status:", self.vendor.status.upper() if self.vendor.status else "N/A"),
            ("Category:", self.vendor.category or "N/A"),
            ("Criticality Score:", str(self.vendor.criticality_score or "N/A")),
            ("Data Classification:", self.vendor.data_classification or "N/A"),
            (
                "Contract Expiry:",
                (
                    self.vendor.contract_expiry.isoformat()
                    if self.vendor.contract_expiry
                    else "N/A"
                ),
            ),
            (
                "Last Assessed:",
                (
                    self.vendor.last_assessed.strftime("%Y-%m-%d")
                    if self.vendor.last_assessed
                    else "N/A"
                ),
            ),
            (
                "Next Assessment Due:",
                (
                    self.vendor.next_assessment_due.isoformat()
                    if self.vendor.next_assessment_due
                    else "N/A"
                ),
            ),
            ("Report Generated:", datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")),
        ]

        for row_idx, (label, value) in enumerate(info_data, 3):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        # Statistics
        stats_row = len(info_data) + 5
        ws.cell(row=stats_row, column=1, value="Statistics").font = Font(
            bold=True, size=14
        )

        stats_data = [
            ("Total Documents:", len(self.documents)),
            ("Total Findings:", len(self.findings)),
            (
                "Critical Findings:",
                sum(1 for f in self.findings if f.severity and f.severity.lower() == "critical"),
            ),
            (
                "High Findings:",
                sum(1 for f in self.findings if f.severity and f.severity.lower() == "high"),
            ),
            ("Total Remediation Tasks:", len(self.remediation_tasks)),
            (
                "Open Tasks:",
                sum(
                    1
                    for t in self.remediation_tasks
                    if t.status not in ["closed", "verified"]
                ),
            ),
            (
                "SLA Breached Tasks:",
                sum(1 for t in self.remediation_tasks if t.sla_breached),
            ),
        ]

        for row_idx, (label, value) in enumerate(stats_data, stats_row + 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        ExcelStyler.auto_size_columns(ws)

    def _add_documents_sheet(self, wb: Workbook) -> None:
        """Add documents sheet."""
        ws = wb.create_sheet("Documents")

        columns = ["ID", "Filename", "Type", "Size (bytes)", "Status", "Uploaded At"]

        for col, header in enumerate(columns, 1):
            ws.cell(row=1, column=col, value=header)

        ExcelStyler.apply_header_style(ws)

        for row_idx, doc in enumerate(self.documents, 2):
            ws.cell(row=row_idx, column=1, value=doc.id)
            ws.cell(row=row_idx, column=2, value=doc.filename)
            ws.cell(row=row_idx, column=3, value=doc.mime_type or "")
            ws.cell(row=row_idx, column=4, value=doc.file_size)
            ws.cell(row=row_idx, column=5, value=doc.processing_status or "")
            ws.cell(
                row=row_idx,
                column=6,
                value=(
                    doc.created_at.strftime("%Y-%m-%d %H:%M")
                    if doc.created_at
                    else ""
                ),
            )

            for col in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER
                ws.cell(row=row_idx, column=col).alignment = CELL_ALIGNMENT

        ExcelStyler.auto_size_columns(ws)
        ExcelStyler.freeze_header_row(ws)

    def _add_findings_sheet(self, wb: Workbook) -> None:
        """Add findings sheet."""
        ws = wb.create_sheet("Findings")

        columns = [
            "Title",
            "Severity",
            "Status",
            "Framework",
            "Control",
            "Description",
            "Remediation",
            "Confidence",
        ]

        for col, header in enumerate(columns, 1):
            ws.cell(row=1, column=col, value=header)

        ExcelStyler.apply_header_style(ws)

        for row_idx, finding in enumerate(self.findings, 2):
            ws.cell(row=row_idx, column=1, value=finding.title)

            severity_cell = ws.cell(
                row=row_idx, column=2, value=finding.severity.upper() if finding.severity else ""
            )
            if finding.severity:
                ExcelStyler.apply_severity_color(severity_cell, finding.severity)

            status_cell = ws.cell(row=row_idx, column=3, value=finding.status.upper() if finding.status else "")
            if finding.status:
                ExcelStyler.apply_status_color(status_cell, finding.status)

            ws.cell(row=row_idx, column=4, value=finding.framework or "")
            ws.cell(row=row_idx, column=5, value=finding.framework_control or "")
            ws.cell(row=row_idx, column=6, value=finding.description)
            ws.cell(row=row_idx, column=7, value=finding.remediation or "")
            ws.cell(
                row=row_idx,
                column=8,
                value=(
                    f"{finding.confidence_score:.2f}"
                    if finding.confidence_score
                    else ""
                ),
            )

            for col in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGNMENT

        ExcelStyler.auto_size_columns(ws)
        ExcelStyler.freeze_header_row(ws)

    def _add_remediation_sheet(self, wb: Workbook) -> None:
        """Add remediation tasks sheet."""
        ws = wb.create_sheet("Remediation")

        columns = [
            "Title",
            "Status",
            "Priority",
            "Assignee",
            "Due Date",
            "SLA Breached",
            "Resolution Notes",
        ]

        for col, header in enumerate(columns, 1):
            ws.cell(row=1, column=col, value=header)

        ExcelStyler.apply_header_style(ws)

        for row_idx, task in enumerate(self.remediation_tasks, 2):
            ws.cell(row=row_idx, column=1, value=task.title)
            ws.cell(row=row_idx, column=2, value=task.status.upper() if task.status else "")

            priority_cell = ws.cell(
                row=row_idx, column=3, value=task.priority.upper() if task.priority else ""
            )
            if task.priority:
                ExcelStyler.apply_priority_color(priority_cell, task.priority)

            ws.cell(
                row=row_idx, column=4, value=task.assignee.email if task.assignee else ""
            )
            ws.cell(
                row=row_idx,
                column=5,
                value=task.due_date.strftime("%Y-%m-%d") if task.due_date else "",
            )

            sla_cell = ws.cell(
                row=row_idx, column=6, value="Yes" if task.sla_breached else "No"
            )
            if task.sla_breached:
                sla_cell.fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
                sla_cell.font = Font(color="FFFFFF", bold=True)

            ws.cell(row=row_idx, column=7, value=task.resolution_notes or "")

            for col in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGNMENT

        ExcelStyler.auto_size_columns(ws)
        ExcelStyler.freeze_header_row(ws)

    async def generate(self) -> bytes:
        """Generate comprehensive vendor report."""
        # Load data from database
        await self._load_data()

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Add sheets
        self._add_overview_sheet(wb)
        self._add_documents_sheet(wb)
        self._add_findings_sheet(wb)
        self._add_remediation_sheet(wb)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
